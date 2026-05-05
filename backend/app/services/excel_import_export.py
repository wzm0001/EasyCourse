from io import BytesIO
from typing import Optional

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.basic_data import (
    Grade, Class_, Course, GradeCourse, Teacher, TeacherCourse,
    Classroom, ClassroomCourse, TeachingArrangement,
)
from app.repositories.basic_data import (
    GradeRepository, ClassRepository, CourseRepository, GradeCourseRepository,
    TeacherRepository, TeacherCourseRepository, ClassroomRepository,
    ClassroomCourseRepository, TeachingArrangementRepository,
)

TEMPLATES = {
    "grades": ["序号", "年级名称", "排序"],
    "classes": ["序号", "年级名称", "班级名称", "班主任姓名"],
    "courses": ["序号", "课程代码", "课程名称", "课程类型(必修/选修)", "每周课时"],
    "teachers": ["序号", "工号", "姓名", "性别", "联系电话", "教研组"],
    "classrooms": ["序号", "教室名称", "类型(普通教室/实验室/专用教室)", "容量"],
    "teaching_arrangements": ["序号", "教师姓名", "课程名称", "班级名称", "每周课时"],
}


async def download_template(data_type: str) -> bytes:
    if data_type not in TEMPLATES:
        raise ValueError(f"不支持的数据类型: {data_type}")
    wb = Workbook()
    ws = wb.active
    ws.title = data_type
    ws.append(TEMPLATES[data_type])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


async def import_data(data_type: str, file: UploadFile, school_id: str, semester_id: str, db: AsyncSession) -> dict:
    if data_type not in TEMPLATES:
        raise ValueError(f"不支持的数据类型: {data_type}")

    content = await file.read()
    wb = load_workbook(BytesIO(content))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    success_count = 0
    errors = []

    if data_type == "grades":
        grade_repo = GradeRepository(db)
        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 2 or not row[1]:
                    errors.append({"row": idx, "error": "年级名称不能为空"})
                    continue
                name = str(row[1]).strip()
                sort_order = int(row[2]) if len(row) > 2 and row[2] is not None else 0
                grade = Grade(
                    school_id=school_id,
                    semester_id=semester_id,
                    name=name,
                    sort_order=sort_order,
                )
                await grade_repo.create(grade)
                success_count += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

    elif data_type == "classes":
        class_repo = ClassRepository(db)
        grade_repo = GradeRepository(db)
        teacher_repo = TeacherRepository(db)
        grades, _ = await grade_repo.get_by_semester(school_id, semester_id, page=1, page_size=1000)
        grade_map = {g.name: g.id for g in grades}
        teachers, _ = await teacher_repo.get_by_semester(school_id, semester_id, page=1, page_size=1000)
        teacher_map = {t.name: t.id for t in teachers}
        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 3 or not row[1] or not row[2]:
                    errors.append({"row": idx, "error": "年级名称和班级名称不能为空"})
                    continue
                grade_name = str(row[1]).strip()
                class_name = str(row[2]).strip()
                teacher_name = str(row[3]).strip() if len(row) > 3 and row[3] else None
                grade_id = grade_map.get(grade_name)
                if not grade_id:
                    errors.append({"row": idx, "error": f"年级 '{grade_name}' 不存在"})
                    continue
                teacher_id = teacher_map.get(teacher_name) if teacher_name else None
                cls = Class_(
                    school_id=school_id,
                    semester_id=semester_id,
                    grade_id=grade_id,
                    name=class_name,
                    teacher_id=teacher_id,
                )
                await class_repo.create(cls)
                success_count += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

    elif data_type == "courses":
        course_repo = CourseRepository(db)
        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 3 or not row[1] or not row[2]:
                    errors.append({"row": idx, "error": "课程代码和课程名称不能为空"})
                    continue
                code = str(row[1]).strip()
                name = str(row[2]).strip()
                course_type = "required"
                if len(row) > 3 and row[3]:
                    ct = str(row[3]).strip()
                    if "选修" in ct:
                        course_type = "elective"
                weekly_hours = int(row[4]) if len(row) > 4 and row[4] is not None else 1
                course = Course(
                    school_id=school_id,
                    code=code,
                    name=name,
                    course_type=course_type,
                    weekly_hours=weekly_hours,
                )
                await course_repo.create(course)
                success_count += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

    elif data_type == "teachers":
        teacher_repo = TeacherRepository(db)
        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 3 or not row[1] or not row[2]:
                    errors.append({"row": idx, "error": "工号和姓名不能为空"})
                    continue
                employee_id = str(row[1]).strip()
                name = str(row[2]).strip()
                gender = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                phone = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                teaching_group = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                existing = await teacher_repo.get_by_employee_id(employee_id, school_id)
                if existing:
                    errors.append({"row": idx, "error": f"工号 '{employee_id}' 已存在"})
                    continue
                teacher = Teacher(
                    school_id=school_id,
                    semester_id=semester_id,
                    name=name,
                    employee_id=employee_id,
                    gender=gender,
                    phone=phone,
                    teaching_group=teaching_group,
                )
                await teacher_repo.create(teacher)
                success_count += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

    elif data_type == "classrooms":
        classroom_repo = ClassroomRepository(db)
        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 2 or not row[1]:
                    errors.append({"row": idx, "error": "教室名称不能为空"})
                    continue
                name = str(row[1]).strip()
                room_type = "normal"
                if len(row) > 2 and row[2]:
                    rt = str(row[2]).strip()
                    if "实验" in rt:
                        room_type = "lab"
                    elif "专用" in rt:
                        room_type = "special"
                capacity = int(row[3]) if len(row) > 3 and row[3] is not None else 50
                classroom = Classroom(
                    school_id=school_id,
                    semester_id=semester_id,
                    name=name,
                    room_type=room_type,
                    capacity=capacity,
                )
                await classroom_repo.create(classroom)
                success_count += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

    elif data_type == "teaching_arrangements":
        arrangement_repo = TeachingArrangementRepository(db)
        teacher_repo = TeacherRepository(db)
        course_repo = CourseRepository(db)
        class_repo = ClassRepository(db)
        teachers, _ = await teacher_repo.get_by_semester(school_id, semester_id, page=1, page_size=1000)
        teacher_map = {t.name: t.id for t in teachers}
        courses, _ = await course_repo.get_by_school(school_id, page=1, page_size=1000)
        course_map = {c.name: c.id for c in courses}
        classes, _ = await class_repo.get_by_semester(school_id, semester_id, page=1, page_size=1000)
        class_map = {c.name: c.id for c in classes}
        for idx, row in enumerate(rows, start=2):
            try:
                if not row or len(row) < 4 or not row[1] or not row[2] or not row[3]:
                    errors.append({"row": idx, "error": "教师姓名、课程名称和班级名称不能为空"})
                    continue
                teacher_name = str(row[1]).strip()
                course_name = str(row[2]).strip()
                class_name = str(row[3]).strip()
                weekly_hours = int(row[4]) if len(row) > 4 and row[4] is not None else 1
                teacher_id = teacher_map.get(teacher_name)
                if not teacher_id:
                    errors.append({"row": idx, "error": f"教师 '{teacher_name}' 不存在"})
                    continue
                course_id = course_map.get(course_name)
                if not course_id:
                    errors.append({"row": idx, "error": f"课程 '{course_name}' 不存在"})
                    continue
                class_id = class_map.get(class_name)
                if not class_id:
                    errors.append({"row": idx, "error": f"班级 '{class_name}' 不存在"})
                    continue
                arrangement = TeachingArrangement(
                    school_id=school_id,
                    semester_id=semester_id,
                    teacher_id=teacher_id,
                    course_id=course_id,
                    class_id=class_id,
                    weekly_hours=weekly_hours,
                )
                await arrangement_repo.create(arrangement)
                success_count += 1
            except Exception as e:
                errors.append({"row": idx, "error": str(e)})

    return {"success_count": success_count, "errors": errors}


async def export_data(data_type: str, school_id: str, semester_id: str, db: AsyncSession) -> bytes:
    if data_type not in TEMPLATES:
        raise ValueError(f"不支持的数据类型: {data_type}")

    wb = Workbook()
    ws = wb.active
    ws.title = data_type
    ws.append(TEMPLATES[data_type])

    if data_type == "grades":
        grade_repo = GradeRepository(db)
        grades, _ = await grade_repo.get_by_semester(school_id, semester_id, page=1, page_size=10000)
        for idx, g in enumerate(grades, start=1):
            ws.append([idx, g.name, g.sort_order])

    elif data_type == "classes":
        class_repo = ClassRepository(db)
        grade_repo = GradeRepository(db)
        teacher_repo = TeacherRepository(db)
        classes, _ = await class_repo.get_by_semester(school_id, semester_id, page=1, page_size=10000)
        for idx, c in enumerate(classes, start=1):
            grade = await grade_repo.get_by_id(c.grade_id)
            grade_name = grade.name if grade else ""
            teacher_name = ""
            if c.teacher_id:
                teacher = await teacher_repo.get_by_id(c.teacher_id)
                teacher_name = teacher.name if teacher else ""
            ws.append([idx, grade_name, c.name, teacher_name])

    elif data_type == "courses":
        course_repo = CourseRepository(db)
        courses, _ = await course_repo.get_by_school(school_id, page=1, page_size=10000)
        for idx, c in enumerate(courses, start=1):
            ct = "必修" if c.course_type.value == "required" else "选修"
            ws.append([idx, c.code, c.name, ct, c.weekly_hours])

    elif data_type == "teachers":
        teacher_repo = TeacherRepository(db)
        teachers, _ = await teacher_repo.get_by_semester(school_id, semester_id, page=1, page_size=10000)
        for idx, t in enumerate(teachers, start=1):
            ws.append([idx, t.employee_id, t.name, t.gender, t.phone, t.teaching_group])

    elif data_type == "classrooms":
        classroom_repo = ClassroomRepository(db)
        classrooms, _ = await classroom_repo.get_by_semester(school_id, semester_id, page=1, page_size=10000)
        for idx, c in enumerate(classrooms, start=1):
            rt_map = {"normal": "普通教室", "lab": "实验室", "special": "专用教室"}
            rt = rt_map.get(c.room_type, c.room_type)
            ws.append([idx, c.name, rt, c.capacity])

    elif data_type == "teaching_arrangements":
        arrangement_repo = TeachingArrangementRepository(db)
        teacher_repo = TeacherRepository(db)
        course_repo = CourseRepository(db)
        class_repo = ClassRepository(db)
        arrangements, _ = await arrangement_repo.get_by_semester(school_id, semester_id, page=1, page_size=10000)
        for idx, a in enumerate(arrangements, start=1):
            teacher = await teacher_repo.get_by_id(a.teacher_id)
            teacher_name = teacher.name if teacher else ""
            course = await course_repo.get_by_id(a.course_id)
            course_name = course.name if course else ""
            cls = await class_repo.get_by_id(a.class_id)
            class_name = cls.name if cls else ""
            ws.append([idx, teacher_name, course_name, class_name, a.weekly_hours])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
