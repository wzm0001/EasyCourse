import io
import zipfile
from typing import Optional, List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.schedule import get_class_schedule, get_teacher_schedule, get_grade_schedule, get_school_schedule
from app.repositories.basic_data import ClassRepository, TeacherRepository
from app.schemas.schedule import ScheduleGrid, ScheduleCellInfo

PERIOD_TYPE_LABELS = {
    "morning_reading": "早读",
    "regular": "正课",
    "evening_study": "晚自习",
}

DAY_LABELS = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]


def _period_label(period_type: str, period_index: int) -> str:
    prefix = PERIOD_TYPE_LABELS.get(period_type, period_type)
    return f"{prefix}{period_index}"


def _cell_content(cell: ScheduleCellInfo, include_classroom: bool = False) -> str:
    parts = []
    if cell.course_name:
        parts.append(cell.course_name)
    if cell.teacher_name:
        parts.append(f"({cell.teacher_name})")
    if include_classroom and cell.classroom_name:
        parts.append(f"[{cell.classroom_name}]")
    return "".join(parts) if parts else ""


def _build_grid_map(cells: List[ScheduleCellInfo]) -> dict:
    grid = {}
    for cell in cells:
        key = (cell.day_of_week, cell.period_type, cell.period_index)
        grid[key] = cell
    return grid


def _get_sorted_periods(cells: List[ScheduleCellInfo]) -> list:
    period_order = {"morning_reading": 0, "regular": 1, "evening_study": 2}
    periods = set()
    for cell in cells:
        periods.add((cell.period_type, cell.period_index))
    sorted_periods = sorted(periods, key=lambda p: (period_order.get(p[0], 99), p[1]))
    return sorted_periods


def _excel_export(grid: ScheduleGrid, include_classroom: bool = False) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = grid.class_name or "课表"

    header_font = Font(bold=True, size=11)
    cell_font = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    ws.cell(row=1, column=1, value="时间段").font = header_font_white
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = thin_border

    for day_idx in range(7):
        col = day_idx + 2
        ws.cell(row=1, column=col, value=DAY_LABELS[day_idx]).font = header_font_white
        ws.cell(row=1, column=col).fill = header_fill
        ws.cell(row=1, column=col).alignment = center_align
        ws.cell(row=1, column=col).border = thin_border

    periods = _get_sorted_periods(grid.cells)
    grid_map = _build_grid_map(grid.cells)

    for row_idx, (period_type, period_index) in enumerate(periods):
        row = row_idx + 2
        label = _period_label(period_type, period_index)
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=1).alignment = center_align
        ws.cell(row=row, column=1).border = thin_border

        for day_idx in range(7):
            col = day_idx + 2
            key = (day_idx + 1, period_type, period_index)
            cell_info = grid_map.get(key)
            content = _cell_content(cell_info, include_classroom) if cell_info else ""
            ws.cell(row=row, column=col, value=content).font = cell_font
            ws.cell(row=row, column=col).alignment = center_align
            ws.cell(row=row, column=col).border = thin_border

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, 9):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_export(grid: ScheduleGrid, include_classroom: bool = False) -> bytes:
    buf = io.BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(buf, pagesize=page_size, leftMargin=15 * mm, rightMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Title"], fontSize=16, spaceAfter=10)
    elements = []

    title = Paragraph(grid.class_name or "课表", title_style)
    elements.append(title)
    elements.append(Spacer(1, 5 * mm))

    periods = _get_sorted_periods(grid.cells)
    grid_map = _build_grid_map(grid.cells)

    num_cols = 8
    table_data = [["时间段"] + DAY_LABELS]

    for period_type, period_index in periods:
        label = _period_label(period_type, period_index)
        row = [label]
        for day_idx in range(7):
            key = (day_idx + 1, period_type, period_index)
            cell_info = grid_map.get(key)
            content = _cell_content(cell_info, include_classroom) if cell_info else ""
            row.append(content)
        table_data.append(row)

    if len(table_data) <= 1:
        table_data.append([""] * num_cols)

    available_width = page_size[0] - 30 * mm
    col_widths = [available_width * 0.1] + [available_width * 0.9 / 7] * 7

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitespace),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))

    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


async def export_class_schedule_excel(class_id: str, semester_id: str, db: AsyncSession) -> bytes:
    grid = await get_class_schedule(class_id, semester_id, db)
    return _excel_export(grid)


async def export_class_schedule_pdf(class_id: str, semester_id: str, db: AsyncSession) -> bytes:
    grid = await get_class_schedule(class_id, semester_id, db)
    return _pdf_export(grid)


async def export_teacher_schedule_excel(teacher_id: str, semester_id: str, db: AsyncSession) -> bytes:
    grid = await get_teacher_schedule(teacher_id, semester_id, db)
    return _excel_export(grid)


async def export_teacher_schedule_pdf(teacher_id: str, semester_id: str, db: AsyncSession) -> bytes:
    grid = await get_teacher_schedule(teacher_id, semester_id, db)
    return _pdf_export(grid)


async def batch_export_teacher_schedules(school_id: str, semester_id: str, db: AsyncSession) -> bytes:
    teacher_repo = TeacherRepository(db)
    teachers, _ = await teacher_repo.get_by_semester(school_id, semester_id, page=1, page_size=10000)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for teacher in teachers:
            grid = await get_teacher_schedule(teacher.id, semester_id, db)
            excel_data = _excel_export(grid)
            safe_name = teacher.name or teacher.id
            zf.writestr(f"{safe_name}.xlsx", excel_data)

    return buf.getvalue()


async def batch_export_class_schedules(school_id: str, semester_id: str, grade_id: Optional[str], db: AsyncSession) -> bytes:
    class_repo = ClassRepository(db)

    if grade_id:
        classes, _ = await class_repo.get_by_grade(grade_id)
    else:
        classes, _ = await class_repo.get_by_semester(school_id, semester_id, page=1, page_size=10000)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for cls in classes:
            grid = await get_class_schedule(cls.id, semester_id, db)
            excel_data = _excel_export(grid)
            safe_name = cls.name or cls.id
            zf.writestr(f"{safe_name}.xlsx", excel_data)

    return buf.getvalue()


async def custom_export(school_id: str, semester_id: str, options: dict, db: AsyncSession) -> bytes:
    fmt = options.get("format", "excel")
    scope = options.get("scope", "class")
    ids = options.get("ids", [])
    include_classroom = options.get("include_classroom", False)

    grids = []
    if scope == "class":
        for class_id in ids:
            grid = await get_class_schedule(class_id, semester_id, db)
            grids.append(grid)
    elif scope == "teacher":
        for teacher_id in ids:
            grid = await get_teacher_schedule(teacher_id, semester_id, db)
            grids.append(grid)
    elif scope == "grade":
        for grade_id in ids:
            grade_grids = await get_grade_schedule(grade_id, semester_id, db)
            grids.extend(grade_grids)

    if fmt == "pdf":
        if len(grids) == 1:
            return _pdf_export(grids[0], include_classroom)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for grid in grids:
                pdf_data = _pdf_export(grid, include_classroom)
                safe_name = grid.class_name or grid.class_id
                zf.writestr(f"{safe_name}.pdf", pdf_data)
        return buf.getvalue()

    if len(grids) == 1:
        return _excel_export(grids[0], include_classroom)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for grid in grids:
            excel_data = _excel_export(grid, include_classroom)
            safe_name = grid.class_name or grid.class_id
            zf.writestr(f"{safe_name}.xlsx", excel_data)
    return buf.getvalue()
